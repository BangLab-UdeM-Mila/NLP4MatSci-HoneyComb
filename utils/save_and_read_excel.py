from csv import excel
import os
import string
from networkx import rescale_layout_dict
import pandas as pd
from openpyxl import load_workbook


def create_parent_directories(path):
    directory = os.path.dirname(path)
    if directory and not os.path.exists(directory):
        os.makedirs(directory)
        print(f"Directory '{directory}' created successfully.")


def sanitize(input_data):
    if isinstance(input_data, str):
        return ''.join(filter(lambda x: x in string.printable, input_data))
    return input_data


def save_to_excel(excel_path, column_values, column_names):
    column_values = [sanitize(value) for value in column_values]
    assert len(column_values) == len(column_names), "column_values and column_names are not the same length"
    
    create_parent_directories(excel_path)
    
    if not os.path.exists(excel_path):
        new_row = pd.DataFrame([column_values], columns=column_names)
        new_row.to_excel(excel_path, index=False)
    else:
        # Load existing workbook
        wb = load_workbook(excel_path)
        ws = wb.active

        # Insert a new row below the header (at position 2)
        ws.insert_rows(2)
        
        # Write the new row data
        for col_num, value in enumerate(column_values, start=1):
            ws.cell(row=2, column=col_num, value=str(value))  # Ensure all values are converted to strings
        
        # Save the workbook
        wb.save(excel_path)


def read_recent_recalled_or_retrieved_documents(excel_path):
    # Read only the first row which is the most recent data (after the header)
    df = pd.read_excel(excel_path, nrows=1)
    assert len(df.columns) == 4, f"Expected 4 columns in {excel_path}, but found {len(df.columns)}"
    return tuple(df.iloc[0])


def save_to_excel_sciq_baseline(excel_path, item, predicted_answer, generated_response):
    question, distractors, correct_answer = item
    data = [question, ", ".join(distractors), correct_answer, predicted_answer, generated_response]
    columns = ["Question", "Distractors", "Correct Answer", "Predicted Answer", "Generated Response"]
    save_to_excel(excel_path, data, columns)


def save_to_excel_mascqa_baseline(excel_path, item, predicted_answer, generated_response):
    qid, question, qtype, correct_answer, topic = item
    data = [qid, qtype, question, generated_response, topic, correct_answer]
    columns = ["Question Info", "Question Type", "Question", "Generated Response", "Topic", "Correct Answer"]
    save_to_excel(excel_path, data, columns)

def save_to_excel_glass_baseline(excel_path, item, predicted_answer, generated_response):
    index, question, correct_answer = item
    data = [index, question, correct_answer, predicted_answer, generated_response]
    columns = ["Index", "Question", "Correct Answer", "Predicted Answer", "Generated Response"]
    save_to_excel(excel_path, data, columns)
